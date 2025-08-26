# app.py
from datetime import datetime, date
from zoneinfo import ZoneInfo
import re
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st
from PIL import Image

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

st.set_page_config(page_title="Enrollment Formatter", layout="centered")

# ----------------------------
# Header / UI
# ----------------------------
try:
    logo = Image.open("header_logo.png")
    st.image(logo, width=300)
except Exception:
    pass

st.title("HCHSP Enrollment Checklist Formatter (2025–2026)")
st.markdown("Upload your **Enrollment.xlsx** file and the **VF funded** file to receive a formatted version with correct class names.")

uploaded_file = st.file_uploader("Upload Enrollment.xlsx", type=["xlsx"], key="enr")
vf_file       = st.file_uploader("Upload VF funded spreadsheet (.xlsx)", type=["xlsx"], key="vf")

# =========================
# Helpers (unchanged + new)
# =========================
def coerce_to_dt(v):
    if pd.isna(v):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return from_excel(v)
        except Exception:
            return None
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except Exception:
                continue
    return None

def most_recent(series):
    dates, texts = [], []
    for v in pd.unique(series.dropna()):
        dt = coerce_to_dt(v)
        if dt:
            dates.append(dt)
        else:
            texts.append(v)
    if dates:
        return max(dates)
    return texts[0] if texts else None

def find_header_row(ws, probe="ST: Participant PID", search_rows=160):
    for row in ws.iter_rows(min_row=1, max_row=search_rows):
        for cell in row:
            if isinstance(cell.value, str) and probe in cell.value:
                return cell.row
    return None

def pid_norm(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace(r"\.0+$", "", regex=True)
    return s.map(lambda x: re.sub(r"\D+", "", x)).str.lstrip("0")

def find_class_columns(cols):
    out = []
    for c in cols:
        if not isinstance(c, str): continue
        low = c.lower().strip()
        if any(bad in low for bad in ["classification", "class size", "capacity"]):
            continue
        if "class name" in low or "classroom" in low or low == "class" or low.startswith("class "):
            out.append(c)
    # keep order, dedupe
    seen=set(); keep=[]
    for c in out:
        if c not in seen:
            seen.add(c); keep.append(c)
    return keep

def find_center_column(cols):
    for name in ["Center Name","Site","Campus","Center","ST: Center Name"]:
        if name in cols:
            return name
    # best-effort: contains 'center'
    for c in cols:
        if isinstance(c, str) and "center" in c.lower():
            return c
    return None

PLACEHOLDER_ANY = re.compile(r"^\s*class(?:room)?\s*\d+\s*$", re.IGNORECASE)
ONLY_NUMERIC    = re.compile(r"^\s*\d+\s*$")

def norm_center(x: str) -> str:
    x = (x or "").lower().strip()
    x = re.sub(r"[\u2013\u2014\-]+", "-", x)
    x = re.sub(r"[^a-z0-9\s\-]", "", x)
    x = re.sub(r"\s+", " ", x)
    for tail in [" isd"," elementary"," elem"," school"]:
        if x.endswith(tail):
            x = x[: -len(tail)]
    return x.strip()

# ---- VF funded parser: Center -> [Class codes] (letters/numbers kept)
CLASS_TOTALS_RE = re.compile(r"^\s*class\s+totals\s*:?\s*$", re.IGNORECASE)
CENTER_LINE_RE  = re.compile(r"^\s*HCHSP\s*--\s*(.+?)\s*$", re.IGNORECASE)
CLASS_LINE_RE   = re.compile(
    r"""^\s*Class\s+
        (?P<code>[A-Za-z0-9][A-Za-z0-9\-\s/]*)   # code (letters/digits, may include hyphen/space/slash)
        (?:\s*[\(:\-]\s*(?P<count>\d{1,3})\s*\)?)?   # optional count; ignored for assignment here
        \s*$""",
    re.IGNORECASE | re.VERBOSE
)
def _clean_code(raw: str) -> str:
    code = re.sub(r"\s*-\s*", "-", raw.strip())
    code = re.sub(r"\s+", "", code)  # E 117A -> E117A
    return code

def parse_vf_funded(vf_xlsx) -> Dict[str, List[str]]:
    """
    Read VF funded and return {VF center text -> [class codes]} (codes like E117, B-114, A01).
    """
    xls = pd.ExcelFile(vf_xlsx)
    sheet = next((s for s in xls.sheet_names if s.lower().startswith("vf_average_funded")), xls.sheet_names[0])
    raw = pd.read_excel(vf_xlsx, sheet_name=sheet, header=None)
    mapping: Dict[str, List[str]] = {}
    current = None
    for val in raw.iloc[:,0].astype(str).fillna(""):
        line = val.strip()
        if not line or CLASS_TOTALS_RE.match(line):
            continue
        m_center = CENTER_LINE_RE.match(line)
        if m_center:
            current = m_center.group(1).strip()
            mapping.setdefault(current, [])
            continue
        m_class = CLASS_LINE_RE.match(line)
        if m_class and current:
            mapping[current].append(_clean_code(m_class.group("code")))
    return mapping

def map_centers(enr_centers: List[str], vf_centers: List[str]) -> Dict[str, str]:
    s_norm = {s: norm_center(s) for s in enr_centers}
    v_norm = {v: norm_center(v) for v in vf_centers}
    rev: Dict[str, List[str]] = {}
    for orig, n in v_norm.items():
        rev.setdefault(n, []).append(orig)
    out = {}
    for s_orig, s_n in s_norm.items():
        if not s_n: continue
        if s_n in rev:
            out[s_orig] = rev[s_n][0]
            continue
        pick = None
        for v_orig, v_n in v_norm.items():
            if s_n in v_n or v_n in s_n:
                pick = v_orig; break
        if pick: out[s_orig] = pick
    return out

def deterministic_assign(block: pd.DataFrame, classes: List[str]) -> pd.Series:
    """
    If a row needs a class (placeholder/empty), assign VF classes in a stable order by PID.
    If Enrollment already has a valid lettered class, we leave it alone.
    """
    if len(classes) == 0 or len(block) == 0:
        return pd.Series([""]*len(block), index=block.index)
    tmp = block.copy()
    tmp["__pid"] = tmp["Participant PID"].astype(str).fillna("")
    tmp = tmp.sort_values("__pid", kind="mergesort")
    out = []
    k = len(classes)
    for i, idx in enumerate(tmp.index):
        out.append(classes[i % k])
    return pd.Series(out, index=tmp.index).reindex(block.index)

# =========================
# Main
# =========================
if uploaded_file:
    # ----------------------------
    # 1) Find the header row
    # ----------------------------
    wb_src = load_workbook(uploaded_file, data_only=True)
    ws_src = wb_src.active

    header_row = find_header_row(ws_src)
    if not header_row:
        st.error("Couldn't find 'ST: Participant PID' in the file.")
        st.stop()
    uploaded_file.seek(0)

    # ----------------------------
    # 2) Load table into pandas
    # ----------------------------
    df = pd.read_excel(uploaded_file, header=header_row - 1)
    df.columns = [c.replace("ST: ", "") if isinstance(c, str) else c for c in df.columns]

    if "Participant PID" not in df.columns:
        st.error("The file is missing 'Participant PID'.")
        st.stop()

    # keep only most-recent row per PID (your original logic)
    df = (
        df.dropna(subset=["Participant PID"])
          .groupby("Participant PID", as_index=False)
          .agg(most_recent)
    )

    # =========================
    # NEW: Override / create class values from VF funded
    # =========================
    if vf_file:
        try:
            vf_map = parse_vf_funded(vf_file)  # {vf_center: [codes]}
        except Exception as e:
            st.warning(f"Could not parse VF funded: {e}")
            vf_map = {}

        center_col = find_center_column(df.columns)
        if center_col is None:
            st.warning("Could not find a Center column in Enrollment; skipping class override.")
        else:
            # normalize a clean center name in Enrollment
            df["__CenterClean"] = df[center_col].astype(str).str.replace(r"^HCHSP\s*--\s*", "", regex=True).str.strip()
            enr_centers = [c for c in df["__CenterClean"].dropna().unique() if c.strip()]
            center_map = map_centers(enr_centers, list(vf_map.keys()))  # Enrollment center -> VF center

            class_cols = find_class_columns(df.columns)
            create_class_col = False
            if not class_cols:
                # create one
                df["Class Name"] = ""
                class_cols = ["Class Name"]
                create_class_col = True

            # For each Enrollment center: replace placeholders or empties with VF classes
            for e_center in enr_centers:
                mask_center = df["__CenterClean"] == e_center
                vf_center = center_map.get(e_center)
                vf_classes = [c for c in (vf_map.get(vf_center, []) if vf_center else []) if c]

                if not vf_classes:
                    # nothing to do for this center
                    continue

                # Pick the column we will write to (first class-like col)
                target_col = class_cols[0]

                # Identify rows that need replacement (placeholder / numeric-only / empty)
                need_mask = mask_center & (
                    df[target_col].astype(str).fillna("").str.strip().eq("") |
                    df[target_col].astype(str).str.match(PLACEHOLDER_ANY) |
                    df[target_col].astype(str).str.match(ONLY_NUMERIC)
                )

                # If we just created the column, assign every row for this center
                if create_class_col:
                    need_mask = mask_center

                block = df.loc[need_mask, ["Participant PID"]].copy()
                if not block.empty:
                    assigned = deterministic_assign(block, vf_classes)
                    df.loc[assigned.index, target_col] = assigned.values

            # drop temp center field
            df.drop(columns=["__CenterClean"], inplace=True, errors="ignore")

    # ----------------------------
    # 3) Write temp workbook
    # ----------------------------
    title_text = "Enrollment Checklist 2025–2026"
    central_now = datetime.now(ZoneInfo("America/Chicago"))
    timestamp_text = central_now.strftime("Generated on %B %d, %Y at %I:%M %p %Z")

    temp_path = "Enrollment_Cleaned.xlsx"
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        # placeholders keep rows allocated
        pd.DataFrame([[title_text]]).to_excel(writer, index=False, header=False, startrow=0)
        pd.DataFrame([[timestamp_text]]).to_excel(writer, index=False, header=False, startrow=1)
        df.to_excel(writer, index=False, startrow=3)

    # ----------------------------
    # 4) Style with openpyxl (unchanged)
    # ----------------------------
    wb = load_workbook(temp_path)
    ws = wb.active

    filter_row = 4               # header row of the data table
    data_start = filter_row + 1  # first data row
    data_end = ws.max_row        # last data row
    max_col = ws.max_column

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(max_col)}{data_end}"

    title_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    ts_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    tcell = ws.cell(row=1, column=1)
    tcell.value = title_text
    tcell.font = Font(size=14, bold=True)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.fill = title_fill

    scell = ws.cell(row=2, column=1)
    scell.value = timestamp_text
    scell.font = Font(size=10, italic=True, color="555555")
    scell.alignment = Alignment(horizontal="center", vertical="center")
    scell.fill = ts_fill

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for cell in ws[filter_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    red_font = Font(color="FF0000", bold=True)

    # Identify key columns
    immun_col = None
    name_col_idx = None
    headers = [ws.cell(row=filter_row, column=c).value for c in range(1, max_col + 1)]
    for idx, h in enumerate(headers, start=1):
        if isinstance(h, str):
            low = h.lower()
            if immun_col is None and "immun" in low:
                immun_col = idx
            if name_col_idx is None and "name" in low:
                name_col_idx = idx
    if name_col_idx is None:
        name_col_idx = 2  # fallback if no "Name" header

    cutoff_date = datetime(2025, 5, 11)
    immun_cutoff = datetime(2024, 5, 11)

    for r in range(data_start, data_end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            cell.border = thin_border

            if val in (None, "", "nan", "NaT"):
                cell.value = "X"
                cell.font = red_font
                continue

            dt = coerce_to_dt(val)
            if dt:
                if c == immun_col and dt < immun_cutoff:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                    cell.font = red_font
                elif dt < cutoff_date:
                    cell.value = "X"
                    cell.font = red_font
                else:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                continue
            # non-date: leave as-is

    width_map = {1: 16, 2: 22}
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = width_map.get(c, 14)

    # ----------------------------
    # 5) Totals at the bottom (unchanged)
    # ----------------------------
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    center = Alignment(horizontal="center", vertical="center")
    for c in range(1, max_col + 1):
        if c <= name_col_idx:
            continue

        valid_count = 0
        for r in range(data_start, data_end + 1):
            if ws.cell(row=r, column=c).value != "X":
                valid_count += 1

        cell = ws.cell(row=total_row, column=c, value=valid_count)
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="thin"))

    # ----------------------------
    # 6) Save and download
    # ----------------------------
    final_output = "Formatted_Enrollment_Checklist.xlsx"
    wb.save(final_output)

    with open(final_output, "rb") as f:
        st.download_button("📥 Download Formatted Excel", f, file_name=final_output)

